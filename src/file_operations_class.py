from abc import ABC, abstractmethod
import json
import openpyxl


class File_worker(ABC):
    """Абстрактный класс для работы с файлами"""

    @abstractmethod
    def save_to_file(self, **kwargs):
        pass

    @abstractmethod
    def read_file(self, **kwargs):
        pass

    @abstractmethod
    def remove_from_file(self, **kwargs):
        pass


class Json_file_worker(File_worker):
    """Класс для работы с Json файлами"""

    def __init__(self, file_path):
        self.file_path = file_path
        self.vacancies = []

    def save_to_file(self, objects_list, file_path):
        """
        Функция сохраняет экземпляр в новый файл.
        :param objects_list: Список объектов
        :param file_path: Название нового файла
        :return:
        """
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump([obj.to_dict() for obj in objects_list], f, ensure_ascii=False, indent=4)
        return print('Файл успешно сохранен')

    # Работает
    def read_file(self):
        """
        Функция чтения данных из файла.
        :return:
        """
        with open(self.file_path, 'r', encoding='utf-8') as file:
            self.vacancies = json.load(file)
        return self.vacancies

    # Работает
    def remove_from_file(self, vacancy_id):
        """
        Функция удаляет вакансию из файла по ключу 'id'
        :param vacancy_id: id вакансии
        :return:
        """
        self.vacancies = [v for v in self.vacancies if v.get('id') != vacancy_id]
        self.save()
        return print('Вакансия успешно удалена.')

    def save(self):
        """
        Функция сохраняет экземпляр в текущий файл.
        :return:
        """
        with open(self.file_path, 'w', encoding='utf-8') as f:
            json.dump(self.vacancies, f, ensure_ascii=False, indent=4)


class Excel_file_worker(File_worker):
    """Класс для работы с Excel файлами"""

    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.append(
            ['id', 'name', 'area', 'salary_min', 'salary_max', 'currency', 'requirement', 'responsibility'])

    # Работает
    def save_to_file(self, data, file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['id', 'name', 'area', 'salary_min', 'salary_max', 'currency', 'requirement', 'responsibility'])
        for vacancy in data:
            sheet.append([vacancy.id, vacancy.name, vacancy.area, vacancy.salary_min,
                          vacancy.salary_max, vacancy.currency, vacancy.requirement, vacancy.responsibility])
        workbook.save(file_path)

    def add_vacancy(self, vacancy):
        vacancy_id = self.sheet.max_row
        self.sheet.append(
            [vacancy_id, vacancy['name'], vacancy['area'], vacancy['salary_min'],
             vacancy['salary_max'], vacancy['currency'], vacancy['requirement'], vacancy['responsibility']]
        )
        self.workbook.save(self.file_path)
        return vacancy_id

    def read_file(self, file_path):
        vacancies = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            vacancy = {
                'id': row[0],
                'name': row[1],
                'area': row[2],
                'salary_min': row[3],
                'salary_max': row[4],
                'currency': row[5],
                'requirement': row[6],
                'responsibility': row[7],
            }
            vacancies.append(vacancy)
        return vacancies

    def remove_from_file(self, vacancy_id):
        for row in self.sheet.iter_rows(min_row=2):
            if row[0].value == vacancy_id:
                self.sheet.delete_rows(row[0].row, amount=1)
                break
        self.workbook.save(self.file_path)
